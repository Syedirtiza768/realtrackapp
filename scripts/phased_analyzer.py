#!/usr/bin/env python3
"""
Phased Spreadsheet Analyzer
────────────────────────────
Phase 1 – Fast structural scan (headers + 100-row sample)
Phase 2 – Global column frequency map & similarity
Phase 3 – Lightweight clustering & provisional classification
Phase 4 – Deep profiling, relationships, quality checks
Deliverables – CSVs + Markdown report
"""

import os, sys, re, json, csv, math, hashlib, warnings, datetime as dt
from pathlib import Path
from collections import Counter, defaultdict
from itertools import combinations
from difflib import SequenceMatcher

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")

# ── Config ────────────────────────────────────────────────────────────────
ROOT = Path(r"D:\apps\listingpro\files")
OUT  = ROOT / "_analysis_outputs"
OUT.mkdir(exist_ok=True)

SAMPLE_ROWS   = 100          # Phase 1 sample
DEEP_ROWS     = 1000         # Phase 4 cap for large files
COL_SIM_THRESH = 0.80        # column-name similarity threshold
CLUSTER_OVERLAP = 0.40       # min Jaccard for clustering

# ── Helpers ───────────────────────────────────────────────────────────────
def plog(msg):
    print(msg, flush=True)

def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()

def normalize_col(name):
    """Lowercase, strip spaces/special chars, collapse underscores."""
    s = re.sub(r"[^a-z0-9]+", "_", str(name).lower().strip())
    return s.strip("_")

def infer_dtype(series):
    """Infer dominant data type from a pandas Series sample."""
    non_null = series.dropna()
    if len(non_null) == 0:
        return "empty"
    sample = non_null.head(200).astype(str)
    counts = Counter()
    for v in sample:
        v = v.strip()
        if v == "":
            continue
        if re.match(r"^-?\d+$", v):
            counts["integer"] += 1
        elif re.match(r"^-?\d+\.\d+$", v):
            counts["decimal"] += 1
        elif re.match(r"^(true|false|yes|no|0|1)$", v, re.I):
            counts["boolean"] += 1
        elif re.match(r"^\d{1,4}[/-]\d{1,2}[/-]\d{1,4}", v):
            counts["date"] += 1
        elif re.match(r"^[\$€£]", v):
            counts["currency"] += 1
        else:
            counts["string"] += 1
    if not counts:
        return "empty"
    top = counts.most_common(1)[0][0]
    total = sum(counts.values())
    if counts[top] / total < 0.6 and len(counts) > 1:
        return "mixed"
    return top

PATTERN_MAP = {
    "sku": re.compile(r"^[A-Z0-9]{3,}[-]?[A-Z0-9]{2,}$", re.I),
    "email": re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.I),
    "phone": re.compile(r"^[\+]?[\d\s\-\(\)]{7,}$"),
    "url": re.compile(r"^https?://", re.I),
    "currency_value": re.compile(r"^[\$€£]\s?\d"),
    "vin": re.compile(r"^[A-HJ-NPR-Z0-9]{17}$", re.I),
    "date_iso": re.compile(r"^\d{4}-\d{2}-\d{2}"),
    "incremental_id": re.compile(r"^\d{4,}$"),
}

def detect_patterns(series):
    """Return dict of pattern_name -> hit_rate for a column sample."""
    non_null = series.dropna().astype(str).head(200)
    if len(non_null) == 0:
        return {}
    hits = Counter()
    for v in non_null:
        v = v.strip()
        for pname, prx in PATTERN_MAP.items():
            if prx.match(v):
                hits[pname] += 1
    n = len(non_null)
    return {k: round(v / n * 100, 1) for k, v in hits.items() if v / n > 0.15}

def jaccard(a, b):
    sa, sb = set(a), set(b)
    if not sa and not sb:
        return 0
    return len(sa & sb) / len(sa | sb)

# ── Phase 1: Fast Structural Scan ────────────────────────────────────────
plog("\n" + "=" * 70)
plog("PHASE 1 — Fast Structural Scan")
plog("=" * 70)

xlsx_files = sorted(ROOT.glob("*.xlsx")) + sorted(ROOT.glob("*.xls"))
csv_files  = sorted(ROOT.glob("*.csv")) + sorted(ROOT.glob("*.tsv"))
# also recurse subfolders (excluding _analysis_outputs)
for sub in ROOT.rglob("*"):
    if "_analysis_outputs" in str(sub):
        continue
    if sub.suffix.lower() in (".xlsx", ".xls") and sub not in xlsx_files:
        xlsx_files.append(sub)
    elif sub.suffix.lower() in (".csv", ".tsv") and sub not in csv_files:
        csv_files.append(sub)

all_files = xlsx_files + csv_files
total = len(all_files)
plog(f"  Total spreadsheet files detected: {total}")

# Phase 1 results: list of dicts
phase1 = []
errors_log = []

for idx, fpath in enumerate(all_files, 1):
    rel = fpath.relative_to(ROOT)
    pct = int(idx / total * 100)
    plog(f"  [{pct:>3}%] ({idx}/{total})  Scanning: {rel}")
    fsize = fpath.stat().st_size
    fmod  = dt.datetime.fromtimestamp(fpath.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    ftype = fpath.suffix.lower()

    try:
        if ftype in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows_iter = ws.iter_rows(max_row=SAMPLE_ROWS + 1)
                sampled = []
                for r in rows_iter:
                    sampled.append([c.value for c in r])
                if not sampled:
                    phase1.append(dict(
                        file=str(rel), sheet=sname, ftype=ftype, fsize=fsize,
                        fmod=fmod, columns=[], col_count=0, sample_rows=0,
                        total_rows=0, header_confidence=0, norm_cols=[]))
                    continue
                headers = [safe_str(h) for h in sampled[0]]
                data = sampled[1:]
                # try to get total row count from openpyxl
                total_rows = ws.max_row if ws.max_row else len(data)
                hdr_conf = 90 if all(isinstance(h, str) and h for h in sampled[0] if h is not None) else 60
                norm = [normalize_col(h) for h in headers]
                phase1.append(dict(
                    file=str(rel), sheet=sname, ftype=ftype, fsize=fsize,
                    fmod=fmod, columns=headers, col_count=len(headers),
                    sample_rows=len(data), total_rows=total_rows,
                    header_confidence=hdr_conf, norm_cols=norm,
                    _data=data, _headers=headers))
            wb.close()
        elif ftype in (".csv", ".tsv"):
            sep = "\t" if ftype == ".tsv" else ","
            df = pd.read_csv(fpath, sep=sep, nrows=SAMPLE_ROWS, encoding="utf-8", on_bad_lines="skip")
            headers = list(df.columns)
            norm = [normalize_col(h) for h in headers]
            phase1.append(dict(
                file=str(rel), sheet="(csv)", ftype=ftype, fsize=fsize,
                fmod=fmod, columns=headers, col_count=len(headers),
                sample_rows=len(df), total_rows=len(df),
                header_confidence=85, norm_cols=norm))
    except Exception as exc:
        errors_log.append(f"  ⚠ {rel}: {exc}")
        plog(f"  ⚠ Error on {rel}: {exc}")

plog(f"\n  ✔ Phase 1 complete — {len(phase1)} sheets profiled from {total} files")
if errors_log:
    plog(f"  ⚠ {len(errors_log)} file/sheet errors logged")

# ── Phase 2: Column Frequency Map ────────────────────────────────────────
plog("\n" + "=" * 70)
plog("PHASE 2 — Common Column Detection")
plog("=" * 70)

col_freq = Counter()          # norm_col -> count of sheets
col_to_files = defaultdict(set)
col_raw_names = defaultdict(set)

for rec in phase1:
    for raw, norm in zip(rec["columns"], rec["norm_cols"]):
        col_freq[norm] += 1
        col_to_files[norm].add(rec["file"])
        col_raw_names[norm].add(raw)

# Merge similar columns
norm_keys = list(col_freq.keys())
merged = {}  # alias -> canonical
for i in range(len(norm_keys)):
    if norm_keys[i] in merged:
        continue
    for j in range(i + 1, len(norm_keys)):
        if norm_keys[j] in merged:
            continue
        if SequenceMatcher(None, norm_keys[i], norm_keys[j]).ratio() >= COL_SIM_THRESH:
            merged[norm_keys[j]] = norm_keys[i]

canonical_freq = Counter()
canonical_files = defaultdict(set)
for k, v in col_freq.items():
    canon = merged.get(k, k)
    canonical_freq[canon] += v
    canonical_files[canon] |= col_to_files[k]

plog(f"  Unique normalised columns  : {len(col_freq)}")
plog(f"  After similarity merging   : {len(canonical_freq)}")
plog(f"  Top-20 recurring columns:")
for col, cnt in canonical_freq.most_common(20):
    raw_variants = col_raw_names.get(col, set())
    # include aliases
    for alias, canon in merged.items():
        if canon == col:
            raw_variants |= col_raw_names.get(alias, set())
    plog(f"    {col:40s}  in {cnt:>3} sheets  variants: {raw_variants}")

# ── Phase 3: Clustering & Provisional Classification ─────────────────────
plog("\n" + "=" * 70)
plog("PHASE 3 — Clustering & Classification")
plog("=" * 70)

# Build normalised column sets per sheet (using canonical names)
sheet_colsets = {}
for rec in phase1:
    key = f"{rec['file']}|{rec['sheet']}"
    normed = set()
    for n in rec["norm_cols"]:
        normed.add(merged.get(n, n))
    sheet_colsets[key] = normed

# Simple agglomerative clustering by Jaccard overlap
keys = list(sheet_colsets.keys())
assigned = {}
clusters = {}
cid = 0
for i in range(len(keys)):
    if keys[i] in assigned:
        continue
    cid += 1
    cluster_members = [keys[i]]
    cluster_cols = set(sheet_colsets[keys[i]])
    assigned[keys[i]] = cid
    for j in range(i + 1, len(keys)):
        if keys[j] in assigned:
            continue
        j_sim = jaccard(cluster_cols, sheet_colsets[keys[j]])
        if j_sim >= CLUSTER_OVERLAP:
            cluster_members.append(keys[j])
            assigned[keys[j]] = cid
            cluster_cols |= sheet_colsets[keys[j]]
    clusters[cid] = dict(members=cluster_members, union_cols=cluster_cols)

plog(f"  Clusters formed: {len(clusters)}")

# Classify each cluster
def classify_cluster(union_cols):
    """Return (label, confidence, evidence) based on column semantics."""
    cols = set(c.lower() for c in union_cols)
    evidence = []
    # Scoring rubric
    product_signals = {"sku","title","brand","price","upc","mpn","description",
                       "item_name","product_name","condition","category",
                       "manufacturer_part_number","part_number","item_sku",
                       "custom_label","ebay_title","quantity"}
    listing_signals = {"ebay_item_id","listing_status","start_price","buy_it_now_price",
                       "listing_type","listing_duration","ebay_category",
                       "item_id","action","site_id","format","duration"}
    fitment_signals = {"make","model","year","engine","submodel","trim",
                       "fitment","vehicle","compatibility"}
    media_signals   = {"image_url","picurl","pic_url","image","photo",
                       "gallery_url","picture_url"}
    inventory_signals = {"quantity","location","warehouse","stock","bin","qty",
                          "available_quantity"}
    financial_signals = {"debit","credit","amount","balance","invoice",
                          "payment","total","subtotal","tax"}

    scores = {}
    for label, sig in [
        ("Entity Table – Product/Part Catalog", product_signals),
        ("Export Format – eBay Listing Template", listing_signals),
        ("Relational Mapping – Vehicle Fitment", fitment_signals),
        ("Media Mapping – Image/URL heavy", media_signals),
        ("Inventory Snapshot – SKU + Quantity", inventory_signals),
        ("Financial Ledger Pattern", financial_signals),
    ]:
        hits = cols & sig
        if hits:
            scores[label] = len(hits)
            evidence.append(f"{label}: matched {hits}")

    if not scores:
        return "Unknown – insufficient signals", 20, ["No strong column matches"]

    best = max(scores, key=scores.get)
    total_possible = max(len(s) for s in [product_signals, listing_signals,
                                            fitment_signals, media_signals,
                                            inventory_signals, financial_signals])
    conf = min(95, int(scores[best] / total_possible * 100) + 40)
    return best, conf, evidence

for cid, cinfo in clusters.items():
    label, conf, ev = classify_cluster(cinfo["union_cols"])
    cinfo["label"] = label
    cinfo["confidence"] = conf
    cinfo["evidence"] = ev
    short_files = set()
    for m in cinfo["members"]:
        short_files.add(m.split("|")[0])
    plog(f"\n  Cluster {cid}: {label}  (confidence {conf}%)")
    plog(f"    Sheets: {len(cinfo['members'])}  Files: {len(short_files)}")
    plog(f"    Key columns (sample): {list(cinfo['union_cols'])[:15]}")
    for e in ev:
        plog(f"    Evidence: {e}")

plog(f"\n  ✔ Phase 3 complete")

# ── Phase 4: Deep Analysis ───────────────────────────────────────────────
plog("\n" + "=" * 70)
plog("PHASE 4 — Deep Profiling & Relationship Detection")
plog("=" * 70)

deep_profiles = []   # per column detail
quality_issues = []
relationship_candidates = []

# Read full data (capped at DEEP_ROWS) for each sheet
processed = 0
for rec in phase1:
    processed += 1
    pct = int(processed / len(phase1) * 100)
    plog(f"  [{pct:>3}%] Deep-profiling: {rec['file']} / {rec['sheet']}")
    fpath = ROOT / rec["file"]
    ftype = rec["ftype"]
    try:
        if ftype in (".xlsx", ".xls"):
            df = pd.read_excel(fpath, sheet_name=rec["sheet"], nrows=DEEP_ROWS,
                               engine="openpyxl")
        else:
            sep = "\t" if ftype == ".tsv" else ","
            df = pd.read_csv(fpath, sep=sep, nrows=DEEP_ROWS,
                             encoding="utf-8", on_bad_lines="skip")
    except Exception as exc:
        errors_log.append(f"  ⚠ Deep read {rec['file']}/{rec['sheet']}: {exc}")
        plog(f"  ⚠ skipped: {exc}")
        continue

    rec["deep_row_count"] = len(df)
    rec["deep_col_count"] = len(df.columns)

    for col in df.columns:
        ser = df[col]
        missing_pct = round(ser.isna().mean() * 100, 1)
        dtype = infer_dtype(ser)
        pats = detect_patterns(ser)
        nuniq = ser.nunique()
        top5 = ser.dropna().astype(str).value_counts().head(5).to_dict() if nuniq > 0 else {}
        uniqueness = round(nuniq / max(len(ser), 1) * 100, 1)

        profile = dict(
            file=rec["file"], sheet=rec["sheet"],
            column=str(col), norm_col=normalize_col(str(col)),
            inferred_type=dtype, missing_pct=missing_pct,
            unique_values=nuniq, uniqueness_pct=uniqueness,
            patterns=pats,
            top5=json.dumps(dict(list(top5.items())[:5]), default=str),
            cluster_id=assigned.get(f"{rec['file']}|{rec['sheet']}", 0)
        )
        deep_profiles.append(profile)

        # Quality checks
        if missing_pct > 70:
            quality_issues.append(f"Null-heavy: {rec['file']}/{rec['sheet']}.{col} — {missing_pct}% missing")
        if dtype == "mixed":
            quality_issues.append(f"Mixed type: {rec['file']}/{rec['sheet']}.{col}")
        if pats.get("currency_value", 0) > 0 and dtype != "currency":
            quality_issues.append(f"Inconsistent currency: {rec['file']}/{rec['sheet']}.{col}")

# Relationship detection: find columns with same normalised name across sheets
plog("\n  Detecting cross-file relationships …")
col_sheet_map = defaultdict(list)  # norm_col -> [(file, sheet, uniqueness)]
for p in deep_profiles:
    col_sheet_map[p["norm_col"]].append((p["file"], p["sheet"], p["uniqueness_pct"]))

pk_candidates = []
fk_pairs = []
for norm_col, locs in col_sheet_map.items():
    if len(locs) < 2:
        continue
    high_uniq = [(f, s, u) for f, s, u in locs if u > 70]
    low_uniq  = [(f, s, u) for f, s, u in locs if u <= 70]
    if high_uniq:
        pk_candidates.append((norm_col, high_uniq))
    if high_uniq and low_uniq:
        for hf, hs, hu in high_uniq:
            for lf, ls, lu in low_uniq:
                fk_pairs.append(dict(
                    column=norm_col,
                    pk_file=hf, pk_sheet=hs, pk_uniqueness=hu,
                    fk_file=lf, fk_sheet=ls, fk_uniqueness=lu))
    # even many-to-many
    if len(high_uniq) >= 2:
        for (f1, s1, u1), (f2, s2, u2) in combinations(high_uniq, 2):
            relationship_candidates.append(dict(
                column=norm_col,
                file_a=f1, sheet_a=s1, uniq_a=u1,
                file_b=f2, sheet_b=s2, uniq_b=u2,
                rel_type="shared_pk_candidate"))

# Join cluster detection
join_clusters = defaultdict(set)
for norm_col, locs in col_sheet_map.items():
    if len(locs) >= 2:
        for f, s, u in locs:
            join_clusters[norm_col].add(f)

plog(f"  Primary-key candidates    : {len(pk_candidates)}")
plog(f"  Foreign-key pairs         : {len(fk_pairs)}")
plog(f"  Cross-file join columns   : {len(join_clusters)}")
plog(f"  Quality issues detected   : {len(quality_issues)}")

# Duplicate detection — fingerprint each sheet by first-5-col hash
plog("  Checking for structurally identical (redundant) files …")
sheet_fingerprints = defaultdict(list)
for rec in phase1:
    colsig = "|".join(sorted(rec["norm_cols"][:20]))
    fp = hashlib.md5(colsig.encode()).hexdigest()
    sheet_fingerprints[fp].append(f"{rec['file']}|{rec['sheet']}")

redundant = {fp: members for fp, members in sheet_fingerprints.items() if len(members) > 1}
if redundant:
    plog(f"  Possible redundant file groups: {len(redundant)}")

plog(f"\n  ✔ Phase 4 complete")

# ── Deliverables ─────────────────────────────────────────────────────────
plog("\n" + "=" * 70)
plog("GENERATING DELIVERABLES")
plog("=" * 70)

# 1) Master Index CSV
mi_path = OUT / "master_index.csv"
with open(mi_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["File","Sheet","FileType","FileSizeBytes","LastModified",
                "Columns","ColumnCount","SampleRows","TotalRows",
                "HeaderConfidence","ClusterID","ClassificationLabel","ClassificationConfidence"])
    for rec in phase1:
        key = f"{rec['file']}|{rec['sheet']}"
        cid_val = assigned.get(key, 0)
        clabel = clusters.get(cid_val, {}).get("label", "")
        cconf  = clusters.get(cid_val, {}).get("confidence", "")
        w.writerow([rec["file"], rec["sheet"], rec["ftype"], rec["fsize"], rec["fmod"],
                    "; ".join(rec["columns"]), rec["col_count"],
                    rec.get("deep_row_count", rec["sample_rows"]),
                    rec["total_rows"], rec["header_confidence"],
                    cid_val, clabel, cconf])
plog(f"  ✔ Master Index → {mi_path}")

# 2) Data Dictionary CSV
dd_path = OUT / "data_dictionary.csv"
with open(dd_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["File","Sheet","Column","NormColumn","InferredType",
                "MissingPct","UniqueValues","UniquenessPct",
                "Patterns","Top5Values","ClusterID"])
    for p in deep_profiles:
        w.writerow([p["file"], p["sheet"], p["column"], p["norm_col"],
                    p["inferred_type"], p["missing_pct"], p["unique_values"],
                    p["uniqueness_pct"], json.dumps(p["patterns"]),
                    p["top5"], p["cluster_id"]])
plog(f"  ✔ Data Dictionary → {dd_path}")

# 3) Column Frequency Map CSV
cf_path = OUT / "column_frequency_map.csv"
with open(cf_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["NormColumn","SheetCount","CanonicalColumn","RawVariants","FilesContaining"])
    for col, cnt in canonical_freq.most_common():
        raw_variants = col_raw_names.get(col, set())
        for alias, canon in merged.items():
            if canon == col:
                raw_variants |= col_raw_names.get(alias, set())
        w.writerow([col, cnt, col, "; ".join(sorted(raw_variants)),
                    "; ".join(sorted(canonical_files[col]))])
plog(f"  ✔ Column Frequency Map → {cf_path}")

# 4) Processing Progress Log
log_path = OUT / "processing_log.txt"
with open(log_path, "w", encoding="utf-8") as f:
    f.write("Processing Log\n" + "=" * 60 + "\n")
    f.write(f"Total files scanned       : {total}\n")
    f.write(f"Total sheets profiled     : {len(phase1)}\n")
    f.write(f"Clusters formed           : {len(clusters)}\n")
    f.write(f"Column profiles generated : {len(deep_profiles)}\n")
    f.write(f"Quality issues            : {len(quality_issues)}\n")
    f.write(f"Errors / skipped          : {len(errors_log)}\n\n")
    if errors_log:
        f.write("Errors:\n")
        for e in errors_log:
            f.write(f"  {e}\n")
    f.write("\nQuality Issues:\n")
    for q in quality_issues:
        f.write(f"  {q}\n")
plog(f"  ✔ Processing Log → {log_path}")

# 5) Full Classification & Relationship Report (Markdown)
rpt_path = OUT / "classification_report.md"
with open(rpt_path, "w", encoding="utf-8") as f:
    f.write("# Spreadsheet Classification & Relationship Report\n\n")
    f.write(f"_Generated: {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}_\n\n")

    # Summary
    f.write("## Summary\n\n")
    f.write(f"| Metric | Value |\n|---|---|\n")
    f.write(f"| Total files | {total} |\n")
    f.write(f"| Total sheets profiled | {len(phase1)} |\n")
    f.write(f"| Clusters | {len(clusters)} |\n")
    f.write(f"| Column profiles | {len(deep_profiles)} |\n")
    f.write(f"| Quality issues | {len(quality_issues)} |\n")
    f.write(f"| Errors/skipped | {len(errors_log)} |\n\n")

    # Cluster groups
    f.write("## Cluster Groups\n\n")
    for cid_k, cinfo in clusters.items():
        f.write(f"### Cluster {cid_k} — {cinfo['label']}  (confidence {cinfo['confidence']}%)\n\n")
        short_files = sorted(set(m.split("|")[0] for m in cinfo["members"]))
        f.write(f"**Files ({len(short_files)}):**\n\n")
        for sf in short_files:
            f.write(f"- `{sf}`\n")
        f.write(f"\n**Sheets:** {len(cinfo['members'])}\n\n")
        f.write(f"**Key columns (union):**\n\n")
        for c in sorted(cinfo["union_cols"]):
            f.write(f"- `{c}`\n")
        f.write(f"\n**Evidence:**\n\n")
        for e in cinfo["evidence"]:
            f.write(f"- {e}\n")
        f.write("\n---\n\n")

    # Structural similarities
    f.write("## Structural Similarities\n\n")
    if redundant:
        f.write("Files with identical column structures (possible duplicates):\n\n")
        for fp, members in redundant.items():
            f.write(f"- Group `{fp[:8]}…`: {', '.join(members)}\n")
        f.write("\n")
    else:
        f.write("No structurally identical files detected.\n\n")

    # Relationship map
    f.write("## Cross-File Relationship Map\n\n")
    f.write("### Join Columns (appear in ≥2 files)\n\n")
    f.write("| Column | Files Sharing |\n|---|---|\n")
    for norm_col, fset in sorted(join_clusters.items(), key=lambda x: -len(x[1])):
        if len(fset) >= 2:
            f.write(f"| `{norm_col}` | {len(fset)} files |\n")
    f.write("\n")

    f.write("### Primary-Key Candidates\n\n")
    f.write("| Column | File/Sheet | Uniqueness % |\n|---|---|---|\n")
    for norm_col, locs in pk_candidates[:30]:
        for fl, sh, u in locs[:5]:
            f.write(f"| `{norm_col}` | {fl} / {sh} | {u}% |\n")
    f.write("\n")

    f.write("### Foreign-Key Candidates\n\n")
    if fk_pairs:
        f.write("| Column | PK File | FK File | PK Uniq | FK Uniq |\n|---|---|---|---|---|\n")
        for fk in fk_pairs[:30]:
            f.write(f"| `{fk['column']}` | {fk['pk_file']} | {fk['fk_file']} | {fk['pk_uniqueness']}% | {fk['fk_uniqueness']}% |\n")
    else:
        f.write("No clear foreign-key relationships detected.\n")
    f.write("\n")

    # Quality
    f.write("## Data Quality & Risk Summary\n\n")
    if quality_issues:
        for q in quality_issues:
            f.write(f"- {q}\n")
    else:
        f.write("No significant quality issues detected.\n")
    f.write("\n")

    # Suggested schema
    f.write("## Suggested Normalised Schema (if detectable)\n\n")
    if pk_candidates and fk_pairs:
        entities = set()
        for norm_col, locs in pk_candidates:
            for fl, sh, u in locs:
                entities.add(f"{fl}/{sh}")
        f.write("Detectable entities (sheets with high-uniqueness identifiers):\n\n")
        for e in sorted(entities):
            f.write(f"- `{e}`\n")
        f.write("\nRelationships:\n\n")
        for fk in fk_pairs[:15]:
            f.write(f"- `{fk['pk_file']}/{fk['pk_sheet']}` → `{fk['fk_file']}/{fk['fk_sheet']}` via `{fk['column']}`\n")
    else:
        f.write("Insufficient strong relationships to propose a normalized schema.\n")

plog(f"  ✔ Classification Report → {rpt_path}")
plog("\n" + "=" * 70)
plog("ALL DONE — deliverables written to: " + str(OUT))
plog("=" * 70)
