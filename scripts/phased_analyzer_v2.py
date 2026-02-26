#!/usr/bin/env python3
"""
Phased Spreadsheet Analyzer v2 — robust, chunked, fail-safe
────────────────────────────────────────────────────────────
Phase 1 – Fast structural scan (headers + 100-row sample via openpyxl)
Phase 2 – Global column frequency map & similarity
Phase 3 – Lightweight clustering & provisional classification
Phase 4 – Deep profiling per column, relationships, quality
Deliverables – master_index.csv, data_dictionary.csv,
              column_frequency_map.csv, processing_log.txt,
              classification_report.md
"""

import os, sys, re, json, csv, math, hashlib, warnings, datetime as dt, traceback
from pathlib import Path
from collections import Counter, defaultdict
from itertools import combinations
from difflib import SequenceMatcher

import openpyxl, pandas as pd

warnings.filterwarnings("ignore")

# ── Config ─────────────────────────────────────────────────────
ROOT = Path(r"D:\apps\listingpro\files")
OUT  = ROOT / "_analysis_outputs"
OUT.mkdir(exist_ok=True)

SAMPLE_ROWS     = 100
DEEP_ROWS       = 1000
COL_SIM_THRESH  = 0.80
CLUSTER_OVERLAP = 0.40

# ── Helpers ────────────────────────────────────────────────────
def plog(msg): print(msg, flush=True)

def safe_str(v):
    if v is None: return ""
    return str(v).strip()

def normalize_col(name):
    s = re.sub(r"[^a-z0-9]+", "_", str(name).lower().strip())
    return s.strip("_")

def infer_dtype(series):
    non_null = series.dropna()
    if len(non_null) == 0: return "empty"
    sample = non_null.head(200).astype(str)
    counts = Counter()
    for v in sample:
        v = v.strip()
        if v == "": continue
        if re.match(r"^-?\d+$", v):                counts["integer"]  += 1
        elif re.match(r"^-?\d+\.\d+$", v):         counts["decimal"]  += 1
        elif re.match(r"^(true|false|yes|no)$",v,re.I): counts["boolean"] += 1
        elif re.match(r"^\d{1,4}[/-]\d{1,2}[/-]\d{1,4}", v): counts["date"] += 1
        elif re.match(r"^[\$€£]", v):              counts["currency"] += 1
        else:                                       counts["string"]   += 1
    if not counts: return "empty"
    top = counts.most_common(1)[0][0]
    total = sum(counts.values())
    if counts[top] / total < 0.6 and len(counts) > 1: return "mixed"
    return top

PATTERN_MAP = {
    "sku":            re.compile(r"^[A-Z0-9]{3,}[-]?[A-Z0-9]{2,}$", re.I),
    "email":          re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.I),
    "phone":          re.compile(r"^[\+]?[\d\s\-\(\)]{7,}$"),
    "url":            re.compile(r"^https?://", re.I),
    "currency_value": re.compile(r"^[\$€£]\s?\d"),
    "vin":            re.compile(r"^[A-HJ-NPR-Z0-9]{17}$", re.I),
    "date_iso":       re.compile(r"^\d{4}-\d{2}-\d{2}"),
    "incremental_id": re.compile(r"^\d{4,}$"),
}

def detect_patterns(series):
    non_null = series.dropna().astype(str).head(200)
    if len(non_null) == 0: return {}
    hits = Counter()
    for v in non_null:
        v = v.strip()
        for pname, prx in PATTERN_MAP.items():
            if prx.match(v): hits[pname] += 1
    n = len(non_null)
    return {k: round(v / n * 100, 1) for k, v in hits.items() if v / n > 0.15}

def jaccard(a, b):
    sa, sb = set(a), set(b)
    if not sa and not sb: return 0
    return len(sa & sb) / len(sa | sb)

# ════════════════════════════════════════════════════════════════
#  PHASE 1 — Fast Structural Scan
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("PHASE 1 — Fast Structural Scan")
plog("=" * 70)

all_files = []
for ext in ("*.xlsx", "*.xls", "*.csv", "*.tsv", "*.ods"):
    for p in ROOT.rglob(ext):
        if "_analysis_outputs" in str(p): continue
        if p not in all_files: all_files.append(p)
all_files.sort()
total = len(all_files)
plog(f"  Total spreadsheet files detected: {total}")

phase1 = []        # list[dict]
errors_log = []

for idx, fpath in enumerate(all_files, 1):
    rel = str(fpath.relative_to(ROOT))
    pct = int(idx / total * 100)
    plog(f"  [{pct:>3}%] ({idx}/{total})  Scanning: {rel}")
    fsize = fpath.stat().st_size
    fmod  = dt.datetime.fromtimestamp(fpath.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    ftype = fpath.suffix.lower()

    try:
        if ftype in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                try:
                    ws = wb[sname]
                    rows_iter = ws.iter_rows(max_row=SAMPLE_ROWS + 1)
                    sampled = []
                    for r in rows_iter:
                        sampled.append([c.value for c in r])
                    if not sampled:
                        phase1.append(dict(
                            file=rel, sheet=sname, ftype=ftype, fsize=fsize,
                            fmod=fmod, columns=[], col_count=0, sample_rows=0,
                            total_rows=0, header_confidence=0, norm_cols=[]))
                        continue
                    headers = [safe_str(h) for h in sampled[0]]
                    data_rows = sampled[1:]
                    total_rows = ws.max_row if ws.max_row else len(data_rows)
                    # header confidence: all-string first row
                    non_empty = [h for h in sampled[0] if h is not None]
                    hdr_conf = 90 if all(isinstance(h, str) for h in non_empty) and len(non_empty) > 0 else 60
                    norm = [normalize_col(h) for h in headers]
                    phase1.append(dict(
                        file=rel, sheet=sname, ftype=ftype, fsize=fsize,
                        fmod=fmod, columns=headers, col_count=len(headers),
                        sample_rows=len(data_rows), total_rows=total_rows,
                        header_confidence=hdr_conf, norm_cols=norm))
                except Exception as se:
                    errors_log.append(f"Sheet error {rel}/{sname}: {se}")
                    plog(f"    ⚠ sheet {sname}: {se}")
            wb.close()
        elif ftype in (".csv", ".tsv"):
            sep = "\t" if ftype == ".tsv" else ","
            df = pd.read_csv(fpath, sep=sep, nrows=SAMPLE_ROWS,
                             encoding="utf-8", on_bad_lines="skip")
            headers = [str(c) for c in df.columns]
            norm = [normalize_col(h) for h in headers]
            phase1.append(dict(
                file=rel, sheet="(csv)", ftype=ftype, fsize=fsize,
                fmod=fmod, columns=headers, col_count=len(headers),
                sample_rows=len(df), total_rows=len(df),
                header_confidence=85, norm_cols=norm))
    except Exception as exc:
        errors_log.append(f"File error {rel}: {exc}")
        plog(f"  ⚠ Error on {rel}: {exc}")

plog(f"\n  ✔ Phase 1 complete — {len(phase1)} sheets from {total} files")
if errors_log:
    plog(f"  ⚠ {len(errors_log)} errors logged")

# ════════════════════════════════════════════════════════════════
#  PHASE 2 — Common Column Detection
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("PHASE 2 — Common Column Detection")
plog("=" * 70)

col_freq       = Counter()               # norm_col → sheet count
col_to_files   = defaultdict(set)
col_raw_names  = defaultdict(set)

for rec in phase1:
    for raw, norm in zip(rec["columns"], rec["norm_cols"]):
        if not norm:
            continue                       # skip empty header columns
        col_freq[norm] += 1
        col_to_files[norm].add(rec["file"])
        col_raw_names[norm].add(raw)

# Merge similar column names
norm_keys = [k for k in col_freq if k]
merged = {}          # alias → canonical
for i in range(len(norm_keys)):
    if norm_keys[i] in merged: continue
    for j in range(i + 1, len(norm_keys)):
        if norm_keys[j] in merged: continue
        if SequenceMatcher(None, norm_keys[i], norm_keys[j]).ratio() >= COL_SIM_THRESH:
            merged[norm_keys[j]] = norm_keys[i]

canonical_freq  = Counter()
canonical_files = defaultdict(set)
for k, v in col_freq.items():
    canon = merged.get(k, k)
    canonical_freq[canon] += v
    canonical_files[canon] |= col_to_files[k]

plog(f"  Unique normalised columns : {len(col_freq)}")
plog(f"  After similarity merging  : {len(canonical_freq)}")
plog(f"  Top-20 recurring columns:")
for col, cnt in canonical_freq.most_common(20):
    raw_variants = col_raw_names.get(col, set())
    for alias, canon in merged.items():
        if canon == col:
            raw_variants |= col_raw_names.get(alias, set())
    plog(f"    {col:45s} {cnt:>3} sheets   variants: {raw_variants}")

# ════════════════════════════════════════════════════════════════
#  PHASE 3 — Clustering & Provisional Classification
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("PHASE 3 — Clustering & Classification")
plog("=" * 70)

# Build normalised column sets per sheet (skip empty cols)
sheet_colsets = {}
for rec in phase1:
    key = f"{rec['file']}|{rec['sheet']}"
    normed = set()
    for n in rec["norm_cols"]:
        if not n: continue
        normed.add(merged.get(n, n))
    sheet_colsets[key] = normed

# Agglomerative clustering by Jaccard overlap
keys = list(sheet_colsets.keys())
assigned = {}
clusters = {}
cid = 0
for i in range(len(keys)):
    if keys[i] in assigned: continue
    cid += 1
    members = [keys[i]]
    ccols = set(sheet_colsets[keys[i]])
    assigned[keys[i]] = cid
    for j in range(i + 1, len(keys)):
        if keys[j] in assigned: continue
        if jaccard(ccols, sheet_colsets[keys[j]]) >= CLUSTER_OVERLAP:
            members.append(keys[j])
            assigned[keys[j]] = cid
            ccols |= sheet_colsets[keys[j]]
    clusters[cid] = dict(members=members, union_cols=ccols)

plog(f"  Clusters formed: {len(clusters)}")

# ── classification rules (evidence-based) ──
def classify_cluster(union_cols):
    cols = set(c.lower() for c in union_cols)
    evidence = []
    signals = {
        "Entity Table – Product/Part Catalog":
            {"sku","title","brand","price","upc","mpn","description","item_name",
             "product_name","condition","category","manufacturer_part_number",
             "part_number","item_sku","custom_label","ebay_title","quantity",
             "oem_number","parts_description","part_title","additional_details",
             "oem","real_price"},
        "Export Format – eBay Listing Template":
            {"ebay_item_id","listing_status","start_price","buy_it_now_price",
             "listing_type","listing_duration","ebay_category","item_id",
             "action","site_id","format","duration","dont_change_this",
             "general_instructions","shippingpolicynames","returnpolicynames",
             "paymentpolicynames"},
        "Reference Lookup – Category Mapping":
            {"category_id","category_name","header","mandatory","values"},
        "Relational Mapping – Vehicle Fitment":
            {"make","model","year","engine","submodel","trim","fitment",
             "vehicle","compatibility","maker"},
        "Media Mapping – Image/URL heavy":
            {"image_url","picurl","pic_url","image","photo","gallery_url",
             "picture_url","image_count","folder_link","image_links_one_per_column_f_g_h",
             "exists_in_drive"},
        "Inventory Snapshot – SKU + Quantity":
            {"quantity","location","warehouse","stock","bin","qty",
             "available_quantity","q"},
        "Financial Ledger Pattern":
            {"debit","credit","amount","balance","invoice","payment","total",
             "subtotal","tax"},
        "Metadata / Instructions Sheet":
            {"info","created","indicates_missing_required_fields",
             "indicates_missing_field_that_will_be_required_soon",
             "fitment_parts_accessories_help"},
    }
    scores = {}
    for label, sig in signals.items():
        hits = cols & sig
        if hits:
            scores[label] = (len(hits), hits)
            evidence.append(f"{label}: matched {hits}")
    if not scores:
        return "Unknown – insufficient signals", 20, ["No strong column matches"]
    best = max(scores, key=lambda k: scores[k][0])
    hit_count = scores[best][0]
    conf = min(95, int(hit_count / max(len(s) for s in signals.values()) * 100) + 40)
    # boost if multiple labels hit — pick highest
    return best, conf, evidence

for cid_k, cinfo in clusters.items():
    label, conf, ev = classify_cluster(cinfo["union_cols"])
    cinfo["label"] = label
    cinfo["confidence"] = conf
    cinfo["evidence"] = ev
    short_files = sorted(set(m.split("|")[0] for m in cinfo["members"]))
    plog(f"\n  Cluster {cid_k}: {label}  (confidence {conf}%)")
    plog(f"    Sheets: {len(cinfo['members'])}   Files: {len(short_files)}")
    plog(f"    Key columns: {sorted(cinfo['union_cols'])[:12]}")
    for e in ev[:3]:
        plog(f"    → {e}")

plog(f"\n  ✔ Phase 3 complete")

# ════════════════════════════════════════════════════════════════
#  PHASE 4 — Deep Analysis (per-column profiling)
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("PHASE 4 — Deep Profiling & Relationship Detection")
plog("=" * 70)

deep_profiles  = []
quality_issues = []
processed      = 0

for rec in phase1:
    processed += 1
    pct = int(processed / len(phase1) * 100)
    if processed % 10 == 0 or processed == len(phase1):
        plog(f"  [{pct:>3}%] Deep-profiling ({processed}/{len(phase1)})")
    try:
        fpath_full = ROOT / rec["file"]
        ftype = rec["ftype"]
        if ftype in (".xlsx", ".xls"):
            df = pd.read_excel(fpath_full, sheet_name=rec["sheet"],
                               nrows=DEEP_ROWS, engine="openpyxl")
        else:
            sep = "\t" if ftype == ".tsv" else ","
            df = pd.read_csv(fpath_full, sep=sep, nrows=DEEP_ROWS,
                             encoding="utf-8", on_bad_lines="skip")
    except Exception as exc:
        errors_log.append(f"Deep-read {rec['file']}/{rec['sheet']}: {exc}")
        continue

    rec["deep_row_count"] = len(df)
    rec["deep_col_count"] = len(df.columns)

    for col in df.columns:
        try:
            ser = df[col]
            missing_pct = round(ser.isna().mean() * 100, 1)
            dtype = infer_dtype(ser)
            pats  = detect_patterns(ser)
            nuniq = ser.nunique()
            top5  = (ser.dropna().astype(str).value_counts()
                     .head(5).to_dict()) if nuniq > 0 else {}
            uniqueness = round(nuniq / max(len(ser), 1) * 100, 1)

            deep_profiles.append(dict(
                file=rec["file"], sheet=rec["sheet"],
                column=str(col), norm_col=normalize_col(str(col)),
                inferred_type=dtype, missing_pct=missing_pct,
                unique_values=nuniq, uniqueness_pct=uniqueness,
                patterns=pats,
                top5=json.dumps(dict(list(top5.items())[:5]), default=str),
                cluster_id=assigned.get(f"{rec['file']}|{rec['sheet']}", 0)
            ))

            if missing_pct > 70:
                quality_issues.append(
                    f"Null-heavy: {rec['file']}/{rec['sheet']}.{col} — {missing_pct}% missing")
            if dtype == "mixed":
                quality_issues.append(
                    f"Mixed type: {rec['file']}/{rec['sheet']}.{col}")
        except Exception as ce:
            errors_log.append(f"Column error {rec['file']}/{rec['sheet']}.{col}: {ce}")

plog(f"  Column profiles: {len(deep_profiles)}")

# ── Relationship detection ──
plog("  Detecting cross-file relationships …")
col_sheet_map = defaultdict(list)        # norm_col → [(file, sheet, uniq%)]
for p in deep_profiles:
    nc = p["norm_col"]
    if not nc: continue
    col_sheet_map[nc].append((p["file"], p["sheet"], p["uniqueness_pct"]))

pk_candidates = []
fk_pairs      = []
for nc, locs in col_sheet_map.items():
    if len(locs) < 2: continue
    high = [(f,s,u) for f,s,u in locs if u > 70]
    low  = [(f,s,u) for f,s,u in locs if u <= 70]
    if high:
        pk_candidates.append((nc, high))
    if high and low:
        for hf,hs,hu in high:
            for lf,ls,lu in low:
                fk_pairs.append(dict(column=nc,pk_file=hf,pk_sheet=hs,
                                     pk_uniq=hu,fk_file=lf,fk_sheet=ls,fk_uniq=lu))

join_clusters = defaultdict(set)
for nc, locs in col_sheet_map.items():
    if len(locs) >= 2:
        for f,s,u in locs:
            join_clusters[nc].add(f)

# Duplicate / redundant detection
sheet_fingerprints = defaultdict(list)
for rec in phase1:
    sig = "|".join(sorted(rec["norm_cols"][:20]))
    fp = hashlib.md5(sig.encode()).hexdigest()
    sheet_fingerprints[fp].append(f"{rec['file']}|{rec['sheet']}")
redundant = {fp: m for fp, m in sheet_fingerprints.items() if len(m) > 1}

plog(f"  PK candidates       : {len(pk_candidates)}")
plog(f"  FK pairs            : {len(fk_pairs)}")
plog(f"  Join columns        : {len(join_clusters)}")
plog(f"  Quality issues      : {len(quality_issues)}")
plog(f"  Redundant groups    : {len(redundant)}")
plog(f"\n  ✔ Phase 4 complete")

# ════════════════════════════════════════════════════════════════
#  DELIVERABLES
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("GENERATING DELIVERABLES")
plog("=" * 70)

# 1) Master Index CSV
mi = OUT / "master_index.csv"
with open(mi, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["File","Sheet","FileType","FileSizeBytes","LastModified",
                "Columns","ColumnCount","SampleRows","TotalRows",
                "HeaderConfidence","ClusterID","Classification","Confidence"])
    for rec in phase1:
        key = f"{rec['file']}|{rec['sheet']}"
        c = assigned.get(key, 0)
        cl = clusters.get(c, {})
        w.writerow([rec["file"], rec["sheet"], rec["ftype"], rec["fsize"],
                    rec["fmod"], "; ".join(rec["columns"]), rec["col_count"],
                    rec.get("deep_row_count", rec["sample_rows"]),
                    rec["total_rows"], rec["header_confidence"],
                    c, cl.get("label",""), cl.get("confidence","")])
plog(f"  ✔ {mi.name}  ({mi.stat().st_size:,} bytes)")

# 2) Data Dictionary CSV
dd = OUT / "data_dictionary.csv"
with open(dd, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["File","Sheet","Column","NormColumn","InferredType",
                "MissingPct","UniqueValues","UniquenessPct",
                "Patterns","Top5Values","ClusterID"])
    for p in deep_profiles:
        w.writerow([p["file"], p["sheet"], p["column"], p["norm_col"],
                    p["inferred_type"], p["missing_pct"], p["unique_values"],
                    p["uniqueness_pct"], json.dumps(p["patterns"]),
                    p["top5"], p["cluster_id"]])
plog(f"  ✔ {dd.name}  ({dd.stat().st_size:,} bytes)")

# 3) Column Frequency Map CSV
cf = OUT / "column_frequency_map.csv"
with open(cf, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["NormColumn","SheetCount","RawVariants","FilesContaining"])
    for col, cnt in canonical_freq.most_common():
        raw = col_raw_names.get(col, set())
        for alias, canon in merged.items():
            if canon == col: raw |= col_raw_names.get(alias, set())
        w.writerow([col, cnt, "; ".join(sorted(raw)),
                    "; ".join(sorted(canonical_files[col]))])
plog(f"  ✔ {cf.name}  ({cf.stat().st_size:,} bytes)")

# 4) Processing Progress Log
lg = OUT / "processing_log.txt"
with open(lg, "w", encoding="utf-8") as f:
    f.write("Processing Log\n" + "=" * 60 + "\n")
    f.write(f"Files scanned          : {total}\n")
    f.write(f"Sheets profiled        : {len(phase1)}\n")
    f.write(f"Clusters formed        : {len(clusters)}\n")
    f.write(f"Column profiles        : {len(deep_profiles)}\n")
    f.write(f"Quality issues         : {len(quality_issues)}\n")
    f.write(f"Errors / skipped       : {len(errors_log)}\n\n")
    if errors_log:
        f.write("─── Errors ───\n")
        for e in errors_log: f.write(f"  {e}\n")
    f.write("\n─── Quality Issues ───\n")
    for q in quality_issues: f.write(f"  {q}\n")
plog(f"  ✔ {lg.name}  ({lg.stat().st_size:,} bytes)")

# 5) Classification & Relationship Report (Markdown)
rpt = OUT / "classification_report.md"
with open(rpt, "w", encoding="utf-8") as f:
    f.write("# Spreadsheet Classification & Relationship Report\n\n")
    f.write(f"_Generated: {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}_\n\n")

    f.write("## Executive Summary\n\n")
    f.write(f"| Metric | Value |\n|---|---|\n")
    f.write(f"| Files scanned | {total} |\n")
    f.write(f"| Sheets profiled | {len(phase1)} |\n")
    f.write(f"| Clusters | {len(clusters)} |\n")
    f.write(f"| Column profiles | {len(deep_profiles)} |\n")
    f.write(f"| Quality issues | {len(quality_issues)} |\n")
    f.write(f"| Errors / skipped | {len(errors_log)} |\n\n")

    # ── Cluster detail ──
    f.write("## Cluster Groups\n\n")
    for cid_k, cinfo in clusters.items():
        sf = sorted(set(m.split("|")[0] for m in cinfo["members"]))
        f.write(f"### Cluster {cid_k} — {cinfo['label']}  (confidence {cinfo['confidence']}%)\n\n")
        f.write(f"**Files ({len(sf)}):**\n\n")
        for s in sf: f.write(f"- `{s}`\n")
        f.write(f"\n**Sheets:** {len(cinfo['members'])}\n\n")
        f.write(f"**Key columns (union, up to 20):**\n\n")
        for c in sorted(cinfo["union_cols"])[:20]: f.write(f"- `{c}`\n")
        f.write(f"\n**Evidence:**\n\n")
        for e in cinfo["evidence"]: f.write(f"- {e}\n")
        f.write("\n---\n\n")

    # ── Structural similarities ──
    f.write("## Structural Similarities / Possible Redundancies\n\n")
    if redundant:
        for fp, members in redundant.items():
            f.write(f"- Fingerprint `{fp[:8]}…`:\n")
            for m in members: f.write(f"  - `{m}`\n")
        f.write("\n")
    else:
        f.write("No structurally identical sheets detected.\n\n")

    # ── Cross-file relationships ──
    f.write("## Cross-File Relationship Map\n\n")
    f.write("### Join Columns (≥2 files)\n\n")
    f.write("| Column | Files |\n|---|---|\n")
    for nc, fset in sorted(join_clusters.items(), key=lambda x: -len(x[1])):
        f.write(f"| `{nc}` | {len(fset)} |\n")
    f.write("\n")

    f.write("### Primary-Key Candidates\n\n")
    f.write("| Column | File / Sheet | Uniqueness |\n|---|---|---|\n")
    for nc, locs in pk_candidates[:40]:
        for fl, sh, u in locs[:3]:
            f.write(f"| `{nc}` | {fl} / {sh} | {u}% |\n")
    f.write("\n")

    f.write("### Foreign-Key Candidates\n\n")
    if fk_pairs:
        f.write("| Column | PK source | FK source | PK uniq | FK uniq |\n|---|---|---|---|---|\n")
        for fk in fk_pairs[:40]:
            f.write(f"| `{fk['column']}` | {fk['pk_file']} | {fk['fk_file']} | {fk['pk_uniq']}% | {fk['fk_uniq']}% |\n")
    else:
        f.write("No clear FK relationships detected.\n")
    f.write("\n")

    # ── Quality ──
    f.write("## Data Quality & Risk Summary\n\n")
    if quality_issues:
        # Categorise
        cats = defaultdict(list)
        for q in quality_issues:
            if q.startswith("Null"):   cats["Null-heavy columns"].append(q)
            elif q.startswith("Mix"):  cats["Mixed-type columns"].append(q)
            else:                      cats["Other"].append(q)
        for cat, items in cats.items():
            f.write(f"### {cat} ({len(items)})\n\n")
            for it in items[:30]: f.write(f"- {it}\n")
            if len(items) > 30: f.write(f"- … and {len(items)-30} more\n")
            f.write("\n")
    else:
        f.write("No significant quality issues detected.\n\n")

    # ── Suggested schema ──
    f.write("## Suggested Normalised Schema\n\n")
    if pk_candidates and fk_pairs:
        entities = set()
        for nc, locs in pk_candidates:
            for fl, sh, u in locs: entities.add(f"{fl}/{sh}")
        f.write("Detectable entities (high-uniqueness identifiers):\n\n")
        for e in sorted(entities)[:30]: f.write(f"- `{e}`\n")
        f.write("\nRelationships:\n\n")
        for fk in fk_pairs[:20]:
            f.write(f"- `{fk['pk_file']}/{fk['pk_sheet']}` → `{fk['fk_file']}/{fk['fk_sheet']}` via `{fk['column']}`\n")
    else:
        f.write("Insufficient strong relationships to propose a normalized schema.\n")

plog(f"  ✔ {rpt.name}  ({rpt.stat().st_size:,} bytes)")

plog("\n" + "=" * 70)
plog(f"ALL DONE — deliverables in: {OUT}")
plog("=" * 70)
