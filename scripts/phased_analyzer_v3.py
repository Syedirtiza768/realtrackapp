#!/usr/bin/env python3
"""
Phased Spreadsheet Analyzer v3 — cleaned, capped, actionable
─────────────────────────────────────────────────────────────
Key improvements over v2:
  • Wide sheets (>300 cols) are flagged and only the top 200 *populated*
    columns are deep-profiled.
  • Quality issues are filtered: only columns with <95% nulls are
    flagged for mixed-type; null-heavy is only reported at sheet level.
  • Clustering uses sheet-type awareness (same sheet name across
    workbooks ⇒ same cluster) plus Jaccard fall-through.
  • Deliverables are compacted and human-readable.
"""

import os, sys, re, json, csv, math, hashlib, warnings, datetime as dt, traceback
from pathlib import Path
from collections import Counter, defaultdict
from itertools import combinations
from difflib import SequenceMatcher

import openpyxl, pandas as pd

warnings.filterwarnings("ignore")

# ── Config ─────────────────────────────────────────────────────
ROOT = Path(r"D:\apps\listingpro\files")
OUT  = ROOT / "_analysis_outputs"
OUT.mkdir(exist_ok=True)

SAMPLE_ROWS      = 100
DEEP_ROWS        = 1000
MAX_PROFILE_COLS = 200      # cap columns profiled per wide sheet
COL_SIM_THRESH   = 0.80
CLUSTER_OVERLAP  = 0.40

# ── Helpers ────────────────────────────────────────────────────
def plog(msg): print(msg, flush=True)
def safe_str(v): return "" if v is None else str(v).strip()

def normalize_col(name):
    s = re.sub(r"[^a-z0-9]+", "_", str(name).lower().strip())
    return s.strip("_")

def infer_dtype(series):
    non_null = series.dropna()
    if len(non_null) == 0: return "empty"
    sample = non_null.head(200).astype(str)
    counts = Counter()
    for v in sample:
        v = v.strip()
        if v == "": continue
        if re.match(r"^-?\d+$", v):                counts["integer"]  += 1
        elif re.match(r"^-?\d+\.\d+$", v):         counts["decimal"]  += 1
        elif re.match(r"^(true|false|yes|no)$",v,re.I): counts["boolean"] += 1
        elif re.match(r"^\d{1,4}[/-]\d{1,2}[/-]\d{1,4}", v): counts["date"] += 1
        elif re.match(r"^[\$€£]", v):              counts["currency"] += 1
        else:                                       counts["string"]   += 1
    if not counts: return "empty"
    top = counts.most_common(1)[0][0]
    total = sum(counts.values())
    if counts[top] / total < 0.6 and len(counts) > 1: return "mixed"
    return top

PATTERN_MAP = {
    "sku":            re.compile(r"^[A-Z0-9]{3,}[-]?[A-Z0-9]{2,}$", re.I),
    "email":          re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.I),
    "phone":          re.compile(r"^[\+]?[\d\s\-\(\)]{7,}$"),
    "url":            re.compile(r"^https?://", re.I),
    "currency_value": re.compile(r"^[\$€£]\s?\d"),
    "vin":            re.compile(r"^[A-HJ-NPR-Z0-9]{17}$", re.I),
    "date_iso":       re.compile(r"^\d{4}-\d{2}-\d{2}"),
    "incremental_id": re.compile(r"^\d{4,}$"),
}

def detect_patterns(series):
    non_null = series.dropna().astype(str).head(200)
    if len(non_null) == 0: return {}
    hits = Counter()
    for v in non_null:
        v = v.strip()
        for pname, prx in PATTERN_MAP.items():
            if prx.match(v): hits[pname] += 1
    n = len(non_null)
    return {k: round(v / n * 100, 1) for k, v in hits.items() if v / n > 0.15}

def jaccard(a, b):
    sa, sb = set(a), set(b)
    if not sa and not sb: return 0
    return len(sa & sb) / len(sa | sb)

# ════════════════════════════════════════════════════════════════
#  PHASE 1 — Fast Structural Scan
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("PHASE 1 — Fast Structural Scan")
plog("=" * 70)

all_files = []
for ext in ("*.xlsx", "*.xls", "*.csv", "*.tsv", "*.ods"):
    for p in ROOT.rglob(ext):
        if "_analysis_outputs" in str(p): continue
        if p not in all_files: all_files.append(p)
all_files.sort()
total_files = len(all_files)
plog(f"  Total spreadsheet files: {total_files}")

phase1 = []
errors_log = []

for idx, fpath in enumerate(all_files, 1):
    rel = str(fpath.relative_to(ROOT))
    pct = int(idx / total_files * 100)
    plog(f"  [{pct:>3}%] ({idx}/{total_files})  {rel}")
    fsize = fpath.stat().st_size
    fmod  = dt.datetime.fromtimestamp(fpath.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    ftype = fpath.suffix.lower()

    try:
        if ftype in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                try:
                    ws = wb[sname]
                    sampled = []
                    for r in ws.iter_rows(max_row=SAMPLE_ROWS + 1):
                        sampled.append([c.value for c in r])
                    if not sampled:
                        phase1.append(dict(file=rel, sheet=sname, ftype=ftype,
                            fsize=fsize, fmod=fmod, columns=[], col_count=0,
                            sample_rows=0, total_rows=0, header_confidence=0,
                            norm_cols=[], is_wide=False))
                        continue
                    headers = [safe_str(h) for h in sampled[0]]
                    total_rows = ws.max_row or len(sampled) - 1
                    non_empty = [h for h in sampled[0] if h is not None]
                    hdr_conf = 90 if all(isinstance(h, str) for h in non_empty) and non_empty else 60
                    norm = [normalize_col(h) for h in headers]
                    is_wide = len(headers) > 300
                    phase1.append(dict(file=rel, sheet=sname, ftype=ftype,
                        fsize=fsize, fmod=fmod, columns=headers,
                        col_count=len(headers), sample_rows=len(sampled)-1,
                        total_rows=total_rows, header_confidence=hdr_conf,
                        norm_cols=norm, is_wide=is_wide))
                except Exception as se:
                    errors_log.append(f"Sheet {rel}/{sname}: {se}")
            wb.close()
        elif ftype in (".csv", ".tsv"):
            sep = "\t" if ftype == ".tsv" else ","
            df = pd.read_csv(fpath, sep=sep, nrows=SAMPLE_ROWS,
                             encoding="utf-8", on_bad_lines="skip")
            headers = [str(c) for c in df.columns]
            norm = [normalize_col(h) for h in headers]
            phase1.append(dict(file=rel, sheet="(csv)", ftype=ftype,
                fsize=fsize, fmod=fmod, columns=headers,
                col_count=len(headers), sample_rows=len(df),
                total_rows=len(df), header_confidence=85,
                norm_cols=norm, is_wide=len(headers)>300))
    except Exception as exc:
        errors_log.append(f"File {rel}: {exc}")
        plog(f"  ⚠ {rel}: {exc}")

plog(f"\n  ✔ Phase 1: {len(phase1)} sheets from {total_files} files")
wide_count = sum(1 for r in phase1 if r.get("is_wide"))
if wide_count:
    plog(f"  ⚠ {wide_count} wide sheets detected (>300 cols) — will cap profiling at {MAX_PROFILE_COLS}")

# ════════════════════════════════════════════════════════════════
#  PHASE 2 — Common Column Detection
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("PHASE 2 — Common Column Detection")
plog("=" * 70)

col_freq       = Counter()
col_to_files   = defaultdict(set)
col_raw_names  = defaultdict(set)

for rec in phase1:
    for raw, norm in zip(rec["columns"], rec["norm_cols"]):
        if not norm: continue
        col_freq[norm] += 1
        col_to_files[norm].add(rec["file"])
        col_raw_names[norm].add(raw)

# Merge similar
norm_keys = [k for k in col_freq if k]
merged = {}
for i in range(len(norm_keys)):
    if norm_keys[i] in merged: continue
    for j in range(i+1, len(norm_keys)):
        if norm_keys[j] in merged: continue
        if SequenceMatcher(None, norm_keys[i], norm_keys[j]).ratio() >= COL_SIM_THRESH:
            merged[norm_keys[j]] = norm_keys[i]

canonical_freq  = Counter()
canonical_files = defaultdict(set)
for k, v in col_freq.items():
    canon = merged.get(k, k)
    canonical_freq[canon] += v
    canonical_files[canon] |= col_to_files[k]

plog(f"  Unique columns        : {len(col_freq)}")
plog(f"  After merge           : {len(canonical_freq)}")
plog(f"  Top-25 recurring columns:")
for col, cnt in canonical_freq.most_common(25):
    raw = col_raw_names.get(col, set())
    for alias, canon in merged.items():
        if canon == col: raw |= col_raw_names.get(alias, set())
    files_n = len(canonical_files[col])
    plog(f"    {col:50s} {cnt:>4} sheets / {files_n:>3} files   {raw}")

# ════════════════════════════════════════════════════════════════
#  PHASE 3 — Clustering & Classification
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("PHASE 3 — Clustering & Classification")
plog("=" * 70)

# ── First pass: group by sheet name (structural identity across workbooks) ──
sheet_name_groups = defaultdict(list)
for rec in phase1:
    key = f"{rec['file']}|{rec['sheet']}"
    sheet_name_groups[rec["sheet"]].append(key)

# Build canonical column-sets per key
sheet_colsets = {}
for rec in phase1:
    key = f"{rec['file']}|{rec['sheet']}"
    normed = set()
    for n in rec["norm_cols"]:
        if not n: continue
        normed.add(merged.get(n, n))
    sheet_colsets[key] = normed

# ── Clustering: sheets with identical sheet-names that share columns
#    are grouped first; then remaining singletons go through Jaccard. ──
assigned = {}
clusters = {}
cid = 0

# Group by sheet name where ≥2 workbooks share it and Jaccard is decent
for sname, members in sheet_name_groups.items():
    if len(members) < 2: continue
    # check pairwise Jaccard of first pair
    representative = sheet_colsets.get(members[0], set())
    if not representative: continue
    group = []
    union_cols = set()
    for m in members:
        ms = sheet_colsets.get(m, set())
        if jaccard(representative, ms) >= 0.30 or ms == representative:
            group.append(m)
            union_cols |= ms
    if len(group) >= 2:
        cid += 1
        for g in group:
            assigned[g] = cid
        clusters[cid] = dict(members=group, union_cols=union_cols, via="sheet-name")

# Remaining un-assigned → Jaccard clustering
remaining = [k for k in sheet_colsets if k not in assigned]
for i in range(len(remaining)):
    k = remaining[i]
    if k in assigned: continue
    cid += 1
    members = [k]
    ccols = set(sheet_colsets[k])
    assigned[k] = cid
    for j in range(i+1, len(remaining)):
        kj = remaining[j]
        if kj in assigned: continue
        if jaccard(ccols, sheet_colsets[kj]) >= CLUSTER_OVERLAP:
            members.append(kj)
            assigned[kj] = cid
            ccols |= sheet_colsets[kj]
    clusters[cid] = dict(members=members, union_cols=ccols, via="jaccard")

plog(f"  Clusters formed: {len(clusters)}")

# ── Classification ──
SIGNALS = {
    "Entity Table – Product/Part Catalog":
        {"sku","title","brand","price","upc","mpn","description","item_name",
         "product_name","condition","category","manufacturer_part_number",
         "part_number","item_sku","custom_label","ebay_title","quantity",
         "oem_number","parts_description","part_title","additional_details",
         "oem","real_price"},
    "Export Format – eBay Listing Template":
        {"action","site_id","format","duration","start_price","buy_it_now_price",
         "listing_type","listing_duration","ebay_category","ebay_item_id",
         "listing_status","item_id","dont_change_this","general_instructions",
         "shippingpolicynames","returnpolicynames","paymentpolicynames"},
    "Reference Lookup – Category Mapping":
        {"category_id","category_name","header","mandatory","values"},
    "Reference Lookup – eBay Item Aspects":
        {"c_brand","c_manufacturer_part_number","c_placement_on_vehicle",
         "c_fitment_type","c_warranty","c_type","c_material",
         "c_surface_finish","c_color"},
    "Relational Mapping – Vehicle Fitment":
        {"make","model","year","engine","submodel","trim","fitment",
         "vehicle","compatibility","maker"},
    "Media Mapping – Image/URL heavy":
        {"image_url","picurl","pic_url","image","photo","gallery_url",
         "picture_url","image_count","folder_link",
         "image_links_one_per_column_f_g_h","exists_in_drive"},
    "Inventory Snapshot – SKU + Quantity":
        {"quantity","location","warehouse","stock","bin","qty",
         "available_quantity","q"},
    "Metadata / Instructions Sheet":
        {"info","created","indicates_missing_required_fields",
         "indicates_missing_field_that_will_be_required_soon",
         "fitment_parts_accessories_help"},
    "Condition Descriptors Reference":
        {"condition_descriptor","conditiondescriptors"},
}

def classify_cluster(union_cols):
    cols = set(c.lower() for c in union_cols)
    evidence = []
    scores = {}
    for label, sig in SIGNALS.items():
        hits = cols & sig
        if hits:
            scores[label] = (len(hits), hits)
            evidence.append(f"{label}: matched {sorted(hits)}")
    if not scores:
        return "Unknown – insufficient signals", 20, ["No strong column matches"]
    best = max(scores, key=lambda k: scores[k][0])
    hit_count = scores[best][0]
    max_sig = max(len(s) for s in SIGNALS.values())
    conf = min(95, int(hit_count / max_sig * 100) + 40)
    return best, conf, evidence

for cid_k, cinfo in clusters.items():
    label, conf, ev = classify_cluster(cinfo["union_cols"])
    cinfo["label"] = label
    cinfo["confidence"] = conf
    cinfo["evidence"] = ev
    short_files = sorted(set(m.split("|")[0] for m in cinfo["members"]))
    plog(f"\n  C{cid_k}: {label}  ({conf}%) — {len(cinfo['members'])} sheets / {len(short_files)} files [{cinfo['via']}]")
    for e in ev[:3]:
        plog(f"      {e}")

plog(f"\n  ✔ Phase 3 complete")

# ════════════════════════════════════════════════════════════════
#  PHASE 4 — Deep Profiling (capped for wide sheets)
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("PHASE 4 — Deep Profiling & Relationships")
plog("=" * 70)

deep_profiles  = []
quality_issues = []
processed      = 0

for rec in phase1:
    processed += 1
    pct = int(processed / len(phase1) * 100)
    if processed % 20 == 0 or processed <= 2 or processed == len(phase1):
        plog(f"  [{pct:>3}%] ({processed}/{len(phase1)})  {rec['file']}/{rec['sheet']}")

    try:
        fpath_full = ROOT / rec["file"]
        ftype = rec["ftype"]
        if ftype in (".xlsx", ".xls"):
            df = pd.read_excel(fpath_full, sheet_name=rec["sheet"],
                               nrows=DEEP_ROWS, engine="openpyxl")
        else:
            sep = "\t" if ftype == ".tsv" else ","
            df = pd.read_csv(fpath_full, sep=sep, nrows=DEEP_ROWS,
                             encoding="utf-8", on_bad_lines="skip")
    except Exception as exc:
        errors_log.append(f"Deep-read {rec['file']}/{rec['sheet']}: {exc}")
        continue

    rec["deep_row_count"] = len(df)
    rec["deep_col_count"] = len(df.columns)

    # For wide sheets: keep only the top MAX_PROFILE_COLS populated columns
    cols_to_profile = list(df.columns)
    if rec.get("is_wide") and len(cols_to_profile) > MAX_PROFILE_COLS:
        # rank by non-null count descending
        pop = df.notna().sum().sort_values(ascending=False)
        cols_to_profile = list(pop.head(MAX_PROFILE_COLS).index)

    for col in cols_to_profile:
        try:
            ser = df[col]
            missing_pct = round(ser.isna().mean() * 100, 1)
            dtype = infer_dtype(ser)
            pats  = detect_patterns(ser)
            nuniq = ser.nunique()
            top5  = (ser.dropna().astype(str).value_counts()
                     .head(5).to_dict()) if nuniq > 0 else {}
            uniqueness = round(nuniq / max(len(ser), 1) * 100, 1)

            deep_profiles.append(dict(
                file=rec["file"], sheet=rec["sheet"],
                column=str(col), norm_col=normalize_col(str(col)),
                inferred_type=dtype, missing_pct=missing_pct,
                unique_values=nuniq, uniqueness_pct=uniqueness,
                patterns=pats,
                top5=json.dumps(dict(list(top5.items())[:5]), default=str),
                cluster_id=assigned.get(f"{rec['file']}|{rec['sheet']}", 0)
            ))

            # Quality flags — only meaningful columns
            if missing_pct > 70 and missing_pct < 100:
                quality_issues.append(
                    f"Null-heavy: {rec['file']}/{rec['sheet']}.{col} ({missing_pct}%)")
            if dtype == "mixed" and missing_pct < 50:
                quality_issues.append(
                    f"Mixed type: {rec['file']}/{rec['sheet']}.{col}")
        except Exception:
            pass

plog(f"  Column profiles generated: {len(deep_profiles)}")
plog(f"  Quality issues logged    : {len(quality_issues)}")

# ── Relationship detection ──
plog("  Cross-file relationship detection …")
col_sheet_map = defaultdict(list)
for p in deep_profiles:
    nc = p["norm_col"]
    if not nc: continue
    col_sheet_map[nc].append((p["file"], p["sheet"], p["uniqueness_pct"]))

pk_candidates = []
fk_pairs      = []
for nc, locs in col_sheet_map.items():
    if len(locs) < 2: continue
    high = [(f,s,u) for f,s,u in locs if u > 70]
    low  = [(f,s,u) for f,s,u in locs if u <= 70]
    if high:
        pk_candidates.append((nc, high))
    if high and low:
        for hf,hs,hu in high:
            for lf,ls,lu in low:
                fk_pairs.append(dict(column=nc,pk_file=hf,pk_sheet=hs,
                                     pk_uniq=hu,fk_file=lf,fk_sheet=ls,fk_uniq=lu))

join_clusters = defaultdict(set)
for nc, locs in col_sheet_map.items():
    if len(locs) >= 2:
        for f,s,u in locs:
            join_clusters[nc].add(f)

# Redundancy
sheet_fps = defaultdict(list)
for rec in phase1:
    sig = "|".join(sorted(set(n for n in rec["norm_cols"] if n))[:30])
    fp = hashlib.md5(sig.encode()).hexdigest()
    sheet_fps[fp].append(f"{rec['file']}|{rec['sheet']}")
redundant = {fp: m for fp, m in sheet_fps.items() if len(m) > 1}

plog(f"  PK candidates     : {len(pk_candidates)}")
plog(f"  FK pairs          : {len(fk_pairs)}")
plog(f"  Join columns      : {len(join_clusters)}")
plog(f"  Redundant groups  : {len(redundant)}")
plog(f"\n  ✔ Phase 4 complete")

# ════════════════════════════════════════════════════════════════
#  DELIVERABLES
# ════════════════════════════════════════════════════════════════
plog("\n" + "=" * 70)
plog("GENERATING DELIVERABLES")
plog("=" * 70)

# 1) Master Index
mi = OUT / "master_index.csv"
with open(mi, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["File","Sheet","Type","SizeBytes","Modified","ColCount","WideFlag",
                "Rows","HeaderConf","ClusterID","Classification","ClassConf"])
    for rec in phase1:
        key = f"{rec['file']}|{rec['sheet']}"
        c = assigned.get(key, 0)
        cl = clusters.get(c, {})
        w.writerow([rec["file"], rec["sheet"], rec["ftype"], rec["fsize"],
                    rec["fmod"], rec["col_count"], rec.get("is_wide",""),
                    rec.get("deep_row_count", rec["sample_rows"]),
                    rec["header_confidence"],
                    c, cl.get("label",""), cl.get("confidence","")])
plog(f"  ✔ master_index.csv          ({mi.stat().st_size:>10,} bytes)")

# 2) Data Dictionary
dd = OUT / "data_dictionary.csv"
with open(dd, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["File","Sheet","Column","NormCol","Type","Missing%",
                "Unique","Uniq%","Patterns","Top5","ClusterID"])
    for p in deep_profiles:
        w.writerow([p["file"], p["sheet"], p["column"], p["norm_col"],
                    p["inferred_type"], p["missing_pct"], p["unique_values"],
                    p["uniqueness_pct"], json.dumps(p["patterns"]),
                    p["top5"], p["cluster_id"]])
plog(f"  ✔ data_dictionary.csv       ({dd.stat().st_size:>10,} bytes)")

# 3) Column Frequency Map
cf = OUT / "column_frequency_map.csv"
with open(cf, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["NormColumn","SheetCount","RawVariants","FileCount"])
    for col, cnt in canonical_freq.most_common():
        raw = col_raw_names.get(col, set())
        for alias, canon in merged.items():
            if canon == col: raw |= col_raw_names.get(alias, set())
        w.writerow([col, cnt, "; ".join(sorted(raw)),
                    len(canonical_files[col])])
plog(f"  ✔ column_frequency_map.csv  ({cf.stat().st_size:>10,} bytes)")

# 4) Processing Log
lg = OUT / "processing_log.txt"
with open(lg, "w", encoding="utf-8") as f:
    f.write("Processing Log\n" + "=" * 60 + "\n")
    f.write(f"Generated       : {dt.datetime.now()}\n")
    f.write(f"Files scanned   : {total_files}\n")
    f.write(f"Sheets profiled : {len(phase1)}\n")
    f.write(f"Wide sheets     : {wide_count}\n")
    f.write(f"Clusters        : {len(clusters)}\n")
    f.write(f"Col profiles    : {len(deep_profiles)}\n")
    f.write(f"Quality issues  : {len(quality_issues)}\n")
    f.write(f"Errors/skipped  : {len(errors_log)}\n\n")
    if errors_log:
        f.write("─── Errors ───\n")
        for e in errors_log: f.write(f"  {e}\n")
        f.write("\n")
    f.write("─── Quality Issues ───\n")
    for q in quality_issues: f.write(f"  {q}\n")
plog(f"  ✔ processing_log.txt        ({lg.stat().st_size:>10,} bytes)")

# 5) Classification & Relationship Report
rpt = OUT / "classification_report.md"
with open(rpt, "w", encoding="utf-8") as md:
    md.write("# Spreadsheet Classification & Relationship Report\n\n")
    md.write(f"_Generated: {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}_\n\n")

    # Summary table
    md.write("## Executive Summary\n\n")
    md.write("| Metric | Value |\n|---|---|\n")
    md.write(f"| Files scanned | {total_files} |\n")
    md.write(f"| Sheets profiled | {len(phase1)} |\n")
    md.write(f"| Wide sheets (>300 cols) | {wide_count} |\n")
    md.write(f"| Clusters | {len(clusters)} |\n")
    md.write(f"| Column profiles | {len(deep_profiles)} |\n")
    md.write(f"| Quality issues | {len(quality_issues)} |\n")
    md.write(f"| Errors / skipped | {len(errors_log)} |\n\n")

    # File inventory overview
    md.write("## File Inventory\n\n")
    md.write("| # | File | Sheets | Total Cols | Largest Sheet Rows | Size KB |\n")
    md.write("|---|---|---|---|---|---|\n")
    file_sheets = defaultdict(list)
    for rec in phase1:
        file_sheets[rec["file"]].append(rec)
    for i, (fname, recs) in enumerate(sorted(file_sheets.items()), 1):
        n_sheets = len(recs)
        total_c = sum(r["col_count"] for r in recs)
        max_rows = max(r.get("deep_row_count", r["sample_rows"]) for r in recs)
        sz = recs[0]["fsize"]
        md.write(f"| {i} | `{fname}` | {n_sheets} | {total_c} | {max_rows} | {sz//1024} |\n")
    md.write("\n")

    # Cluster detail
    md.write("## Cluster Groups\n\n")
    # Sort clusters: named first, then unknowns
    sorted_cids = sorted(clusters.keys(),
        key=lambda c: (0 if "Unknown" not in clusters[c].get("label","") else 1, c))
    for cid_k in sorted_cids:
        cinfo = clusters[cid_k]
        sf = sorted(set(m.split("|")[0] for m in cinfo["members"]))
        sheet_names = sorted(set(m.split("|")[1] for m in cinfo["members"]))
        md.write(f"### Cluster {cid_k} — {cinfo['label']}  ({cinfo['confidence']}%)\n\n")
        md.write(f"- **Method:** {cinfo['via']}\n")
        md.write(f"- **Sheets:** {len(cinfo['members'])} across {len(sf)} files\n")
        md.write(f"- **Sheet names:** {', '.join(sheet_names[:5])}\n")
        md.write(f"- **Key columns:** {', '.join(sorted(cinfo['union_cols'])[:15])}\n\n")
        md.write("**Evidence:**\n\n")
        for e in cinfo["evidence"]: md.write(f"- {e}\n")
        md.write("\n**Files:**\n\n")
        for s in sf[:10]: md.write(f"- `{s}`\n")
        if len(sf) > 10: md.write(f"- … and {len(sf)-10} more\n")
        md.write("\n---\n\n")

    # Structural similarities
    md.write("## Structural Similarities / Redundancies\n\n")
    if redundant:
        md.write(f"Found **{len(redundant)}** groups of sheets with identical column fingerprints:\n\n")
        for fp, members in sorted(redundant.items(), key=lambda x: -len(x[1])):
            md.write(f"**Group** (fingerprint `{fp[:8]}…`, {len(members)} sheets):\n\n")
            for m in members[:8]: md.write(f"- `{m}`\n")
            if len(members) > 8: md.write(f"- … and {len(members)-8} more\n")
            md.write("\n")
    else:
        md.write("No structurally identical sheets.\n\n")

    # Relationships
    md.write("## Cross-File Relationship Map\n\n")
    md.write("### Join Columns (≥ 2 files)\n\n")
    md.write("| Column | # Files | Sample Files |\n|---|---|---|\n")
    for nc, fset in sorted(join_clusters.items(), key=lambda x: -len(x[1])):
        sample_f = ", ".join(sorted(fset)[:3])
        if len(fset) > 3: sample_f += " …"
        md.write(f"| `{nc}` | {len(fset)} | {sample_f} |\n")
    md.write("\n")

    md.write("### Primary-Key Candidates (uniqueness > 70%)\n\n")
    md.write("| Column | File / Sheet | Uniqueness |\n|---|---|---|\n")
    shown = 0
    for nc, locs in pk_candidates:
        for fl, sh, u in locs[:2]:
            md.write(f"| `{nc}` | `{fl}` / {sh} | {u}% |\n")
            shown += 1
        if shown > 50: break
    md.write("\n")

    md.write("### Foreign-Key Candidates\n\n")
    if fk_pairs:
        md.write("| Column | PK source | FK source | PK uniq | FK uniq |\n|---|---|---|---|---|\n")
        for fk in fk_pairs[:40]:
            md.write(f"| `{fk['column']}` | `{fk['pk_file']}` / {fk['pk_sheet']} | `{fk['fk_file']}` / {fk['fk_sheet']} | {fk['pk_uniq']}% | {fk['fk_uniq']}% |\n")
    else:
        md.write("No clear FK relationships detected.\n")
    md.write("\n")

    # Quality
    md.write("## Data Quality & Risk Summary\n\n")
    if quality_issues:
        cats = defaultdict(list)
        for q in quality_issues:
            if q.startswith("Null"):   cats["Null-heavy columns (>70% missing)"].append(q)
            elif q.startswith("Mix"):  cats["Mixed-type columns"].append(q)
            else:                      cats["Other"].append(q)
        for cat, items in cats.items():
            md.write(f"### {cat} ({len(items)})\n\n")
            for it in items[:40]: md.write(f"- {it}\n")
            if len(items) > 40: md.write(f"- _… and {len(items)-40} more_\n")
            md.write("\n")
    else:
        md.write("No significant quality issues.\n\n")

    # Schema suggestion
    md.write("## Suggested Normalised Schema\n\n")
    if pk_candidates:
        md.write("### Detected Entity Tables\n\n")
        for nc, locs in pk_candidates[:15]:
            for fl, sh, u in locs[:1]:
                md.write(f"- **`{nc}`** ({u}% unique) in `{fl}/{sh}`\n")
        md.write("\n")
    if fk_pairs:
        md.write("### Detected Relationships\n\n")
        for fk in fk_pairs[:20]:
            md.write(f"- `{fk['pk_file']}/{fk['pk_sheet']}` →  `{fk['fk_file']}/{fk['fk_sheet']}` via **`{fk['column']}`**\n")
    elif not pk_candidates:
        md.write("Insufficient strong relationships to propose a normalized schema.\n")

plog(f"  ✔ classification_report.md  ({rpt.stat().st_size:>10,} bytes)")

plog("\n" + "=" * 70)
plog(f"ALL DONE — deliverables in: {OUT}")
plog("=" * 70)
